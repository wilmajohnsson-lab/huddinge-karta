#!/usr/bin/env python3
"""
Import Source/*.xlsx into the PocketBase `items` collection.

PocketBase is the source of truth for items. This script is a one-way bulk
importer from the original Excel sources; it dedups by (cat, name, host, addr)
and updates existing PB records in-place. After the import you must press
"Publicera" in the admin UI (or call POST /api/publish/items-json) to push the
resulting items.json to GitHub and trigger a deploy.

Usage:
    # Dry-run: parse Excel, show what would change, do not write to PB
    python3 scripts/import_excel.py

    # Actually write
    PB_PASSWORD='...' python3 scripts/import_excel.py --apply

    # Also delete PB items that match the Excel signature but are not in Excel
    PB_PASSWORD='...' python3 scripts/import_excel.py --apply --prune

Environment variables:
    PB_URL       default http://192.168.86.112:8090  (LAN, bypasses CF WAF)
    PB_EMAIL     default admin@huddinge.mreh.site
    PB_PASSWORD  required for --apply
"""

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime

import openpyxl

# ── Swedish month names ────────────────────────────────────────────────────
SV_MON = {
    1: 'jan', 2: 'feb', 3: 'mar', 4: 'apr', 5: 'maj', 6: 'jun',
    7: 'jul', 8: 'aug', 9: 'sep', 10: 'okt', 11: 'nov', 12: 'dec',
}

# ── Area slug / display-name tables ───────────────────────────────────────
# The 15 canonical kommundelar per Huddinge kommun.
# Source: designer-confirmed list (2026-06).
AREA_SLUG = {
    # canonical 15 (lower-cased input → slug)
    'flemingsberg':        'flemingsberg',
    'skogås':              'skogas',
    'trångsund':           'trangsund',
    'sjödalen-fullersta':  'sjodalen-fullersta',
    'sjödalen':            'sjodalen-fullersta',
    'fullersta':           'sjodalen-fullersta',
    'gladö-lissma':        'glado-lissma',
    'gladö':               'glado-lissma',
    'lissma':              'glado-lissma',
    'stuvsta':             'stuvsta',
    'vårby':               'varby',
    'segeltorp':           'segeltorp',
    'glömsta':             'glomsta',
    'vidja-ågesta':        'vidja-agesta',
    'vidja':               'vidja-agesta',
    'ågesta':              'vidja-agesta',
    'snättringe':          'snattringe',
    'länna':               'lanna',
    'loviseberg':          'loviseberg',
    'kungens kurva':       'kungens-kurva',
    'kungens-kurva':       'kungens-kurva',
    'högmora':             'hogmora',
    # alternative names / sub-areas folded into nearest kommundel
    'huddinge':            'sjodalen-fullersta',
    'huddinge centrum':    'sjodalen-fullersta',
    'vistaberg':           'glomsta',     # Vistaberg ligger inom Glömsta kommundel
    'lida':                'glado-lissma',
}

AREA_DISPLAY = {
    'flemingsberg':       'Flemingsberg',
    'skogas':             'Skogås',
    'trangsund':          'Trångsund',
    'sjodalen-fullersta': 'Sjödalen-Fullersta',
    'glado-lissma':       'Gladö-Lissma',
    'stuvsta':            'Stuvsta',
    'varby':              'Vårby',
    'segeltorp':          'Segeltorp',
    'glomsta':            'Glömsta',
    'vidja-agesta':       'Vidja-Ågesta',
    'snattringe':         'Snättringe',
    'lanna':              'Länna',
    'loviseberg':         'Loviseberg',
    'kungens-kurva':      'Kungens Kurva',
    'hogmora':            'Högmora',
}

DEFAULT_AREA = 'sjodalen-fullersta'

def to_area_id(raw):
    if not raw:
        return DEFAULT_AREA
    s = raw.strip().lower()
    if s in AREA_SLUG:
        return AREA_SLUG[s]
    ascii_s = (s.replace('å','a').replace('ä','a').replace('ö','o')
                 .replace(' ','-').replace('/','-'))
    return ascii_s if ascii_s in AREA_DISPLAY else DEFAULT_AREA

def place_to_area(place_str):
    """Rough area from a place/address string — returns one of the 15 canonical kommundelar."""
    if not place_str:
        return DEFAULT_AREA
    p = place_str.lower()
    # match in priority order: most specific names first
    if 'kungens kurva' in p or 'kungens-kurva' in p or 'ikea' in p:
        return 'kungens-kurva'
    if 'flemingsberg' in p or 'södertörn' in p or 'sh ' in p or 'alfred nobels' in p:
        return 'flemingsberg'
    if 'skogås' in p or 'storvretsvägen' in p:
        return 'skogas'
    if 'trångsund' in p:
        return 'trangsund'
    if 'vårby' in p:
        return 'varby'
    if 'segeltorp' in p or 'juringe' in p:
        return 'segeltorp'
    if 'gladö' in p or 'lissma' in p:
        return 'glado-lissma'
    if 'stuvsta' in p:
        return 'stuvsta'
    if 'snättringe' in p:
        return 'snattringe'
    if 'glömsta' in p or 'vistaberg' in p or 'solhagavägen' in p or 'myrstuguberget' in p:
        return 'glomsta'
    if 'vidja' in p or 'ågesta' in p:
        return 'vidja-agesta'
    if 'länna' in p:
        return 'lanna'
    if 'loviseberg' in p:
        return 'loviseberg'
    if 'högmora' in p:
        return 'hogmora'
    # Huddinge centrum / Sjödalen / Fullersta and unmarked items default here
    return DEFAULT_AREA

def coords_to_area(lat, lng):
    """Fallback: coarse geo-box assignment using the 15 canonical kommundelar.

    These boxes are approximate; place_to_area() (name-based) is preferred
    when a place string is available. Order matters — the first match wins.
    """
    # ── Far east / south-east cluster ────────────────────────────────────
    if lng > 18.10:
        if lat < 59.20:               return 'lanna'
        if lat < 59.225:              return 'trangsund'
        if lat < 59.235:              return 'hogmora'
        return 'skogas'
    # ── Vidja-Ågesta (sydost) ────────────────────────────────────────────
    if lng > 18.05 and lat < 59.21:   return 'vidja-agesta'
    # ── Flemingsberg & Loviseberg (south) ────────────────────────────────
    if lat < 59.205:                  return 'loviseberg'
    if lat < 59.225 and lng < 17.97:  return 'flemingsberg'
    # ── Gladö-Lissma (south rural) ───────────────────────────────────────
    if lat < 59.215 and 17.97 <= lng <= 18.05:  return 'glado-lissma'
    # ── Vårby (NW) ───────────────────────────────────────────────────────
    if lat > 59.275 and lng < 17.92:  return 'varby'
    # ── Kungens Kurva (N, near E4/IKEA) ──────────────────────────────────
    if lat > 59.27 and 17.91 <= lng < 17.95:  return 'kungens-kurva'
    # ── Segeltorp (N) ────────────────────────────────────────────────────
    if lat > 59.265:                  return 'segeltorp'
    # ── Snättringe / Stuvsta (mid-N) ─────────────────────────────────────
    if lat > 59.245 and lng < 17.96:  return 'snattringe'
    if lat > 59.245:                  return 'stuvsta'
    # ── Glömsta (W of Huddinge centrum) ──────────────────────────────────
    if lng < 17.96:                   return 'glomsta'
    # Default: Sjödalen-Fullersta (Huddinge centrum etc.)
    return DEFAULT_AREA

# ── Date helpers ──────────────────────────────────────────────────────────
def format_date_range(dates):
    if not dates:
        return 'Löpande'
    cleaned = sorted(set(
        d.date() if isinstance(d, datetime) else d for d in dates
    ))
    if len(cleaned) == 1:
        d = cleaned[0]
        return f"{d.day} {SV_MON[d.month]}"
    first, last = cleaned[0], cleaned[-1]
    if first.month == last.month:
        return f"{first.day}–{last.day} {SV_MON[first.month]}"
    return f"{first.day} {SV_MON[first.month]}–{last.day} {SV_MON[last.month]}"

def clean_time(t):
    if not t:
        return ''
    t = str(t).strip()
    # "15.30" → "15:30"
    t = re.sub(r'(\d{1,2})\.(\d{2})', r'\1:\2', t)
    # single dash to en-dash in ranges
    t = re.sub(r'(\d{2})-(\d{2})', r'\1–\2', t)
    return t

# ── Category helpers ──────────────────────────────────────────────────────
def event_cat(kategori):
    if not kategori:
        return 'event'
    k = str(kategori).lower()
    if 'musik' in k:
        return 'musik'
    return 'event'

# ── Image pools ───────────────────────────────────────────────────────────
IMG_POOL = {
    'musik': [
        'https://images.unsplash.com/photo-1459749411175-04bf5292ceea?w=600&q=80',
        'https://images.unsplash.com/photo-1415201364774-f6f0bb35f28f?w=600&q=80',
        'https://images.unsplash.com/photo-1493225457124-a3eb161ffa5f?w=600&q=80',
    ],
    'event': [
        'https://images.unsplash.com/photo-1540575467063-178a50c2df87?w=600&q=80',
        'https://images.unsplash.com/photo-1524178232363-1fb2b075b655?w=600&q=80',
        'https://images.unsplash.com/photo-1476820865390-c52aeebb9891?w=600&q=80',
        'https://images.unsplash.com/photo-1531243269054-5ebf6f34081e?w=600&q=80',
        'https://images.unsplash.com/photo-1503095396549-807759245b35?w=600&q=80',
        'https://images.unsplash.com/photo-1486325212027-8081e485255e?w=600&q=80',
    ],
    'konst': [
        'https://images.unsplash.com/photo-1531243269054-5ebf6f34081e?w=600&q=80',
        'https://images.unsplash.com/photo-1518998053901-5348d3961a04?w=600&q=80',
        'https://images.unsplash.com/photo-1541701494587-cb58502866ab?w=600&q=80',
    ],
    'motes': [
        'https://images.unsplash.com/photo-1497366216548-37526070297c?w=600&q=80',
        'https://images.unsplash.com/photo-1481627834876-b7833e8f5570?w=600&q=80',
        'https://images.unsplash.com/photo-1506905925346-21bda4d32df4?w=600&q=80',
    ],
}
_img_idx = {k: 0 for k in IMG_POOL}

def pick_img(cat):
    pool = IMG_POOL[cat]
    img = pool[_img_idx[cat] % len(pool)]
    _img_idx[cat] += 1
    return img

# ── Coordinate lookup ─────────────────────────────────────────────────────
# Static table for well-known Huddinge addresses / landmarks
STATIC_COORDS = {
    'sjödalsparken':                  (59.2354, 17.9813),
    'huddinge centrum':               (59.2372, 17.9820),
    'huddinge torg':                  (59.2367, 17.9811),
    'sjödalstorget':                  (59.2354, 17.9813),
    'paradistorget':                  (59.2350, 17.9809),
    'lilla sjödalstorget':            (59.2363, 17.9827),
    'kommunalvägen 26':               (59.23858, 17.98625),
    'kommunalvägen 27':               (59.23858, 17.98620),
    'kommunalvägen 28a':              (59.23882, 17.98763),
    'kommunalvägen 28':               (59.23882, 17.98763),
    'röntgenvägen 15':                (59.22336, 17.93925),
    'ebba bååts torg 20':             (59.21870, 17.94933),
    'skogåstorget':                   (59.21762, 18.15313),
    'skogåstorget 17':                (59.21762, 18.15313),
    'skogåstorget 7-9':               (59.21762, 18.15313),
    'paradistorget 16':               (59.23508, 17.98091),
    'vårby allé 16':                  (59.26074, 17.87886),
    'storvretsvägen 60':              (59.21863, 18.15361),
    'sjödalstorget 1':                (59.23540, 17.98136),
    'kyrkogårdsvägen 6':              (59.23853, 17.98632),
    'alfred nobels allé 20':          (59.21867, 17.94453),
    'alfred nobels allé':             (59.21829, 17.94188),
    'melodivägen 22':                 (59.23400, 17.97500),
    'visättravägen 13':               (59.21700, 17.95000),
    'mariatorget 7':                  (59.23570, 17.98140),
    'sördalavägen 7':                 (59.23100, 17.99200),
    'magasinet, juringe gård':        (59.27241, 17.92918),
    'juringe gård':                   (59.27241, 17.92918),
    'flemingsberg':                   (59.21936, 17.94203),
}

# Built from aktorlista at runtime
_addr_coords: dict = {}

def _norm_addr(addr: str) -> str:
    if not addr:
        return ''
    # Remove postal code "141 xx City"
    addr = re.sub(r',?\s*\d{3}\s*\d{2}\s*\w[\w\s]*$', '', addr)
    return addr.strip().lower()

def load_aktor_coord_index():
    wb = openpyxl.load_workbook('Source/aktorlista.xlsx')
    ws = wb.active
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 9)]
        forening, plats, stadsdel, kategori, adress, lat, lng, bild = row
        if not (lat and lng):
            continue
        try:
            lat_f, lng_f = float(lat), float(lng)
        except (ValueError, TypeError):
            continue
        for key in filter(None, [
            _norm_addr(adress),
            (plats or '').strip().lower(),
            (forening or '').strip().lower(),
        ]):
            if key:
                _addr_coords[key] = (lat_f, lng_f)

def lookup_coords(adress: str) -> tuple[float, float]:
    if not adress:
        return (59.2372, 17.9820)
    key = adress.strip().lower()

    # 1. Static table (exact)
    if key in STATIC_COORDS:
        return STATIC_COORDS[key]
    # 2. Static table (substring)
    for k, v in STATIC_COORDS.items():
        if k in key or key in k:
            return v
    # 3. Aktorlista index (exact)
    norm = _norm_addr(adress)
    if norm in _addr_coords:
        return _addr_coords[norm]
    # 4. Aktorlista index (substring)
    for k, v in _addr_coords.items():
        if k and len(k) > 4 and (k in norm or norm in k):
            return v
    # 5. Default: Huddinge centrum
    return (59.2372, 17.9820)

# ── Loaders ───────────────────────────────────────────────────────────────

def load_motes(id_start: int) -> list[dict]:
    """aktorlista → motes (venues / organisations with a fixed location)."""
    items = []
    wb = openpyxl.load_workbook('Source/aktorlista.xlsx')
    ws = wb.active
    iid = id_start

    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 9)]
        forening, plats, stadsdel, kategori, adress, lat, lng, bild = row

        name = (plats or forening or '').strip()
        if not name or not lat or not lng:
            continue
        try:
            lat_f, lng_f = float(lat), float(lng)
        except (ValueError, TypeError):
            continue

        host = (forening or name).strip()
        kat   = (str(kategori) if kategori else '').strip()
        area  = to_area_id(stadsdel)
        addr  = (adress or '').strip()
        desc  = kat if kat else 'Kulturplats i Huddinge'

        if len(desc) > 80:
            desc = desc[:77] + '…'

        long_parts = [f'{name} är en kulturplats i {AREA_DISPLAY.get(area, area)}.']
        if kat:
            long_parts.append(f'Typ: {kat}.')
        if addr:
            long_parts.append(f'Adress: {addr}.')
        long_desc = ' '.join(long_parts)

        items.append({
            'id':       iid,
            'cat':      'motes',
            'name':     name,
            'desc':     desc,
            'longDesc': long_desc,
            'date':     'Löpande',
            'time':     'Öppet löpande',
            'loc':      name,
            'addr':     addr,
            'host':     host,
            'area':     area,
            'free':     True,
            'img':      pick_img('motes'),
            'url':      'https://www.huddinge.se',
            'lat':      lat_f,
            'lng':      lng_f,
        })
        iid += 1

    return items


def load_konst(id_start: int) -> list[dict]:
    """konstlista → konst (permanent public art with coordinates)."""
    items = []
    wb = openpyxl.load_workbook('Source/konstlista.xlsx')
    ws = wb.active
    iid = id_start

    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 13)]
        check, namn, konstnär, år, plats, form, utomhus, bild, latitude, longitud, beskrivning, bild2 = row

        if not (namn and latitude and longitud):
            continue
        try:
            lat_f, lng_f = float(latitude), float(longitud)
        except (ValueError, TypeError):
            continue
        # Sanity check: must be in Huddinge area
        if not (59.0 < lat_f < 59.5 and 17.0 < lng_f < 18.5):
            continue

        name    = str(namn).strip()
        artist  = str(konstnär).strip() if konstnär else 'Okänd konstnär'
        year    = str(int(år)) if isinstance(år, (int, float)) and år and int(år) > 0 else ''
        place   = str(plats).strip() if plats else ''
        form_s  = str(form).strip() if form else ''

        desc = f'{form_s} av {artist}' if form_s else f'Konstverk av {artist}'
        if len(desc) > 80:
            desc = desc[:77] + '…'

        if beskrivning and str(beskrivning).strip():
            long_desc = str(beskrivning).strip()
            if year:
                long_desc = f'({year}) {long_desc}'
        else:
            long_desc = f'{desc}.'
            if year:
                long_desc = f'({year}) {long_desc}'
            if place:
                long_desc += f' Belägen vid {place}.'

        # For konst we prefer coordinate-based assignment because Plats names
        # ("Forelltorget", "Sjödalsparken") rarely match a kommundel name. Fall
        # back to place_to_area only when place_to_area finds a real keyword.
        area = place_to_area(place)
        if area == DEFAULT_AREA and lat_f and lng_f:
            area = coords_to_area(lat_f, lng_f)

        items.append({
            'id':       iid,
            'cat':      'konst',
            'name':     name,
            'desc':     desc,
            'longDesc': long_desc,
            'date':     'Permanent',
            'time':     'Dygnet runt',
            'loc':      place or 'Huddinge',
            'addr':     place or 'Huddinge',
            'host':     artist,
            'area':     area,
            'free':     True,
            'img':      pick_img('konst'),
            'url':      'https://www.huddinge.se/uppleva-och-gora/konst-och-kultur/',
            'utomhus':  str(utomhus).strip().lower() == 'ja' if utomhus else False,
            'lat':      lat_f,
            'lng':      lng_f,
        })
        iid += 1

    return items


def load_events(id_start: int) -> list[dict]:
    """eventlista → event / musik items.

    Rows with the same (org, event-name, address) are merged into a
    single item with a collapsed date range.
    """
    wb = openpyxl.load_workbook('Source/eventlista.xlsx')
    ws = wb.active

    # Group rows by (org, name, address)
    groups: dict = {}
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 14)]
        org, namn, typ, adress, koordinater, kommundel, pris, datum, tid, kategori, passar, anmalan, länk = row

        if not namn:
            continue

        key = (
            str(org   or '').strip(),
            str(namn  or '').strip(),
            str(adress or '').strip(),
        )

        if key not in groups:
            groups[key] = {
                'org':        str(org   or '').strip(),
                'name':       str(namn  or '').strip(),
                'type':       str(typ   or '').strip(),
                'addr':       str(adress or '').strip(),
                'kommundel':  str(kommundel or '').strip(),
                'dates':      [],
                'tid':        str(tid   or '').strip() if tid else '',
                'kategori':   str(kategori or '').strip(),
                'passar':     str(passar  or '').strip() if passar else '',
                'free':       pris == 0 or pris is None,
                'pris':       int(pris) if isinstance(pris, (int, float)) else None,
                'anmalan':    str(anmalan or '').strip() if anmalan else '',
                'url':        str(länk  or '').strip() if länk else '',
            }
        g = groups[key]
        if datum:
            g['dates'].append(datum)
        if pris == 0 or pris is None:
            g['free'] = True          # free on any day → mark free
        if länk and not g['url']:
            g['url'] = str(länk).strip()

    items = []
    iid = id_start
    for key, g in groups.items():
        cat        = event_cat(g['kategori'])
        date_str   = format_date_range(g['dates'])
        time_str   = clean_time(g['tid'])
        coords     = lookup_coords(g['addr'])

        # Area: prefer kommundel xlsx column, then address keyword, then coords
        kd = g['kommundel']
        if kd and kd not in ('None', ''):
            area = to_area_id(kd)
        else:
            area = place_to_area(g['addr'])
            if area == DEFAULT_AREA and coords[0] and coords[1]:
                area = coords_to_area(coords[0], coords[1])

        # desc (≤ 80 chars): type · audience
        passar = g['passar'] if g['passar'] not in ('', 'None', 'ingen info finns') else ''
        desc   = g['type'] or g['kategori'] or 'Evenemang'
        if passar:
            candidate = f"{desc} · {passar}"
            desc = candidate if len(candidate) <= 80 else desc
        if len(desc) > 80:
            desc = desc[:77] + '…'

        # longDesc
        sentences = []
        typ = g['type']
        if typ:
            sentences.append(f'{g["name"]} är ett evenemang av typen {typ.lower()}')
        if passar:
            sentences.append(f'passar för {passar}')
        if g['anmalan'] not in ('', 'None', 'nej', 'ingen info finns'):
            sentences.append('anmälan krävs')
        if not g['free']:
            sentences.append('avgiftsbelagd')
        if sentences:
            long_desc = '. '.join(sentences).capitalize() + '.'
        else:
            long_desc = f'Evenemang arrangerat av {g["org"]}.'

        # Derive registration / cta from anmalan + price + link
        anm = g['anmalan'].lower().strip()
        registration = True if anm == 'ja' else (False if anm == 'nej' else None)
        has_link = g['url'] and g['url'] not in ('https://www.huddinge.se', '')
        if registration is True:
            cta = 'apply'   # Anmäl mig
        elif g['pris'] and g['pris'] > 0:
            cta = 'buy'     # Köp biljett
        elif has_link:
            cta = 'info'    # Mer info
        else:
            cta = None

        items.append({
            'id':           iid,
            'cat':          cat,
            'name':         g['name'],
            'desc':         desc,
            'longDesc':     long_desc,
            'date':         date_str,
            'time':         time_str,
            'loc':          g['addr'],
            'addr':         g['addr'],
            'host':         g['org'],
            'area':         area,
            'free':         g['free'],
            'pris':         g['pris'],
            'registration': registration,
            'cta':          cta,
            'cta_url':      g['url'] if has_link else None,
            'img':          pick_img(cat),
            'url':          g['url'] or 'https://www.huddinge.se',
            'lat':          coords[0],
            'lng':          coords[1],
        })
        iid += 1

    return items


# ── PocketBase API helpers ────────────────────────────────────────────────
class PBError(Exception):
    pass

def _http(method: str, url: str, *, token: str | None = None,
          body: dict | None = None, timeout: int = 30) -> dict:
    headers = {'Content-Type': 'application/json'}
    if token:
        headers['Authorization'] = token
    data = json.dumps(body).encode('utf-8') if body is not None else None
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            raw = r.read().decode('utf-8') or '{}'
            return json.loads(raw)
    except urllib.error.HTTPError as e:
        msg = e.read().decode('utf-8', errors='replace')[:400]
        raise PBError(f'{method} {url} → HTTP {e.code}: {msg}') from None
    except urllib.error.URLError as e:
        raise PBError(f'{method} {url} → {e.reason}') from None

def pb_auth(pb_url: str, email: str, password: str) -> str:
    res = _http('POST', f'{pb_url}/api/collections/_superusers/auth-with-password',
                body={'identity': email, 'password': password})
    return res['token']

def pb_list_items(pb_url: str, token: str) -> list[dict]:
    out = []
    page = 1
    while True:
        res = _http('GET',
                    f'{pb_url}/api/collections/items/records?perPage=200&page={page}',
                    token=token)
        out.extend(res.get('items', []))
        if page >= res.get('totalPages', 1):
            break
        page += 1
    return out

def pb_create(pb_url: str, token: str, payload: dict) -> dict:
    return _http('POST', f'{pb_url}/api/collections/items/records',
                 token=token, body=payload)

def pb_update(pb_url: str, token: str, record_id: str, payload: dict) -> dict:
    return _http('PATCH', f'{pb_url}/api/collections/items/records/{record_id}',
                 token=token, body=payload)

def pb_delete(pb_url: str, token: str, record_id: str) -> None:
    _http('DELETE', f'{pb_url}/api/collections/items/records/{record_id}',
          token=token)

# ── Item ↔ PB payload mapping ─────────────────────────────────────────────
# Fields we send to PB. legacy_id is intentionally NOT in this list — the
# publish hook generates sequential ids from PB record creation order.
PB_FIELDS = (
    'name', 'cat', 'desc', 'longDesc', 'date', 'time', 'loc', 'addr', 'host',
    'area', 'free', 'pris', 'registration', 'cta', 'cta_url', 'img', 'url',
    'lat', 'lng',
)

def _reg_to_pb(v) -> str:
    if v is True:  return 'yes'
    if v is False: return 'no'
    return 'unknown'

def _reg_from_pb(v) -> bool | None:
    return {'yes': True, 'no': False}.get(v, None)

def to_pb_payload(item: dict) -> dict:
    """Translate a parsed Excel item into a PB record payload."""
    p = {
        'name':         item['name'],
        'cat':          item['cat'],
        'desc':         item.get('desc') or '',
        'longDesc':     item.get('longDesc') or '',
        'date':         item.get('date') or '',
        'time':         item.get('time') or '',
        'loc':          item.get('loc') or '',
        'addr':         item.get('addr') or '',
        'host':         item.get('host') or '',
        'area':         item['area'],
        'free':         bool(item.get('free')),
        'registration': _reg_to_pb(item.get('registration')),
        'lat':          item['lat'],
        'lng':          item['lng'],
    }
    if item.get('pris') is not None:    p['pris']    = int(item['pris'])
    if item.get('cta'):                 p['cta']     = item['cta']
    if item.get('cta_url'):             p['cta_url'] = item['cta_url']
    if item.get('img'):                 p['img']     = item['img']
    if item.get('url'):                 p['url']     = item['url']
    return p

def sig(item_or_record: dict) -> tuple:
    """Stable identifier for matching Excel items to PB records."""
    return (
        (item_or_record.get('cat')  or '').strip(),
        (item_or_record.get('name') or '').strip(),
        (item_or_record.get('host') or '').strip(),
        (item_or_record.get('addr') or '').strip(),
    )

def diff_payload(payload: dict, record: dict) -> dict:
    """Return only the fields where payload differs from record."""
    out = {}
    for k, v in payload.items():
        rv = record.get(k)
        # Treat None / empty string as equivalent for text fields
        if v == '' and (rv is None or rv == ''):
            continue
        if k in ('lat', 'lng', 'pris') and rv is not None and v is not None:
            try:
                if abs(float(v) - float(rv)) < 1e-9:
                    continue
            except (TypeError, ValueError):
                pass
        if v != rv:
            out[k] = v
    return out

# ── Main ──────────────────────────────────────────────────────────────────
def parse_excel() -> list[dict]:
    print('Loading aktorlista coordinate index…')
    load_aktor_coord_index()

    print('Importing aktorlista  → motes…')
    motes  = load_motes(1)

    print('Importing konstlista  → konst…')
    konst  = load_konst(1 + len(motes))

    print('Importing eventlista  → event / musik…')
    events = load_events(1 + len(motes) + len(konst))

    all_items = motes + konst + events
    print()
    print(f'  {len(motes):>4}  motes    (aktorlista)')
    print(f'  {len(konst):>4}  konst    (konstlista, with coords)')
    print(f'  {len(events):>4}  events   (eventlista, grouped)')
    print(f'  ────')
    print(f'  {len(all_items):>4}  total items parsed')

    # Sanity: warn on coord fallback
    missing_coords = [i for i in all_items
                      if i['lat'] == 59.2372 and i['lng'] == 17.9820]
    if missing_coords:
        print(f'\n  ⚠  {len(missing_coords)} items fell back to Huddinge centrum coords:')
        for i in missing_coords[:10]:
            print(f'     [{i["cat"]}] {i["name"]} — addr: "{i["addr"]}"')
        if len(missing_coords) > 10:
            print(f'     … and {len(missing_coords) - 10} more')

    # Sanity: warn on duplicate signatures inside Excel
    seen: dict[tuple, dict] = {}
    dups = []
    for it in all_items:
        s = sig(it)
        if s in seen:
            dups.append((s, it['name']))
        seen[s] = it
    if dups:
        print(f'\n  ⚠  {len(dups)} duplicate (cat,name,host,addr) signatures inside Excel — '
              f'later rows will overwrite earlier ones in PB:')
        for s, n in dups[:5]:
            print(f'     {s}')

    return all_items

def main() -> int:
    ap = argparse.ArgumentParser(description='Import Excel → PocketBase')
    ap.add_argument('--apply', action='store_true',
                    help='Actually write to PocketBase (default: dry-run)')
    ap.add_argument('--prune', action='store_true',
                    help='Delete PB items not present in current Excel '
                         '(only with --apply)')
    args = ap.parse_args()

    pb_url   = os.environ.get('PB_URL',   'http://192.168.86.112:8090')
    pb_email = os.environ.get('PB_EMAIL', 'admin@huddinge.mreh.site')
    pb_pw    = os.environ.get('PB_PASSWORD', '')

    parsed = parse_excel()

    print()
    print(f'PocketBase: {pb_url} (as {pb_email})')
    if args.apply and not pb_pw:
        print('  ✗ --apply requires PB_PASSWORD env var', file=sys.stderr)
        return 2

    # Always authenticate so we can dry-run a real diff
    if not pb_pw:
        print('  (no PB_PASSWORD set — cannot fetch existing PB state)')
        print('  Set PB_PASSWORD to see a real diff, or pass --apply to write.')
        return 0

    print('  Authenticating…', end=' ', flush=True)
    try:
        token = pb_auth(pb_url, pb_email, pb_pw)
    except PBError as e:
        print('FAIL'); print(f'  ✗ {e}', file=sys.stderr); return 2
    print('ok')

    print('  Fetching existing items…', end=' ', flush=True)
    try:
        existing = pb_list_items(pb_url, token)
    except PBError as e:
        print('FAIL'); print(f'  ✗ {e}', file=sys.stderr); return 2
    print(f'{len(existing)} found')

    by_sig: dict[tuple, dict] = {}
    for r in existing:
        by_sig[sig(r)] = r

    # Dedup parsed list by signature (last row wins, matches the warning above)
    parsed_by_sig: dict[tuple, dict] = {}
    for item in parsed:
        parsed_by_sig[sig(item)] = item

    creates: list[dict] = []
    updates: list[tuple[dict, dict, dict]] = []   # (record, payload, diff)
    unchanged = 0
    for s, item in parsed_by_sig.items():
        payload = to_pb_payload(item)
        if s in by_sig:
            d = diff_payload(payload, by_sig[s])
            if d:
                updates.append((by_sig[s], payload, d))
            else:
                unchanged += 1
        else:
            creates.append(payload)

    orphans = [r for s, r in by_sig.items() if s not in parsed_by_sig]

    print()
    print('=== Diff ===')
    print(f'  create:    {len(creates):>4}')
    print(f'  update:    {len(updates):>4}')
    print(f'  unchanged: {unchanged:>4}')
    print(f'  orphan:    {len(orphans):>4}  '
          '(in PB but not in Excel — would be deleted with --prune)')

    # Show samples
    if creates:
        print('\n  create samples:')
        for p in creates[:5]:
            print(f'    + [{p["cat"]}] {p["name"]}  ({p["host"]})')
        if len(creates) > 5: print(f'    … and {len(creates) - 5} more')
    if updates:
        print('\n  update samples:')
        for r, p, d in updates[:5]:
            keys = ', '.join(sorted(d.keys()))
            print(f'    ~ [{r["cat"]}] {r["name"]}  ({keys})')
        if len(updates) > 5: print(f'    … and {len(updates) - 5} more')
    if orphans:
        print('\n  orphan samples:')
        for r in orphans[:5]:
            print(f'    - [{r["cat"]}] {r["name"]}  ({r["host"]})')
        if len(orphans) > 5: print(f'    … and {len(orphans) - 5} more')

    if not args.apply:
        print('\n(dry-run — re-run with --apply to write)')
        return 0

    # ── Apply ─────────────────────────────────────────────────────────────
    print('\n=== Applying ===')
    ok = err = 0
    for p in creates:
        try:
            pb_create(pb_url, token, p)
            ok += 1
        except PBError as e:
            err += 1
            print(f'  ✗ create [{p["cat"]}] {p["name"]}: {e}', file=sys.stderr)
    print(f'  created:  {ok}/{len(creates)}')

    ok = err = 0
    for r, p, d in updates:
        try:
            pb_update(pb_url, token, r['id'], d)
            ok += 1
        except PBError as e:
            err += 1
            print(f'  ✗ update [{r["cat"]}] {r["name"]}: {e}', file=sys.stderr)
    print(f'  updated:  {ok}/{len(updates)}')

    if args.prune and orphans:
        ok = err = 0
        for r in orphans:
            try:
                pb_delete(pb_url, token, r['id'])
                ok += 1
            except PBError as e:
                err += 1
                print(f'  ✗ delete [{r["cat"]}] {r["name"]}: {e}', file=sys.stderr)
        print(f'  deleted:  {ok}/{len(orphans)}')
    elif orphans:
        print(f'  orphans:  {len(orphans)} kept (no --prune flag)')

    print('\nDone. Open https://huddinge-admin.mreh.site/publicera.html '
          'to publish to the live site.')
    return 0

if __name__ == '__main__':
    sys.exit(main())
